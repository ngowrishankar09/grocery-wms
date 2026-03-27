"""
Shared Spreadsheet router
Endpoints:
  GET    /sheets/workbooks/               → list all workbooks
  POST   /sheets/workbooks/               → create workbook
  DELETE /sheets/workbooks/{wb_id}        → delete workbook
  PATCH  /sheets/workbooks/{wb_id}        → rename workbook

  GET    /sheets/{workbook_id}/           → get full workbook (all sheets + data)
  POST   /sheets/{workbook_id}/           → create a new sheet in workbook
  DELETE /sheets/{workbook_id}/{sheet_id} → delete a sheet
  PATCH  /sheets/{workbook_id}/{sheet_id} → rename sheet / reorder

  POST   /sheets/{workbook_id}/{sheet_id}/columns → add column
  DELETE /sheets/{workbook_id}/{sheet_id}/columns/{col_id} → delete column
  PATCH  /sheets/{workbook_id}/{sheet_id}/columns/{col_id} → rename column / reorder

  POST   /sheets/{workbook_id}/{sheet_id}/rows    → add row(s)
  DELETE /sheets/{workbook_id}/{sheet_id}/rows/{row_id} → delete row
  PATCH  /sheets/{workbook_id}/{sheet_id}/rows/{row_id} → update row colour

  PATCH  /sheets/{workbook_id}/{sheet_id}/cells   → bulk update cells (list of {row_id, col_id, value})
  GET    /sheets/{workbook_id}/poll               → last_updated timestamp for live sync

  GET    /sheets/{workbook_id}/{sheet_id}/export  → export sheet as Excel
"""

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from pydantic import BaseModel
from typing import Optional, List, Any
from datetime import datetime
import json, io

import sys, os
sys.path.append(os.path.dirname(os.path.dirname(__file__)))
from database import get_db
from models import (
    SpreadsheetWorkbook, SpreadsheetSheet,
    SpreadsheetColumn, SpreadsheetRow, SpreadsheetCell
)
from security import get_current_user, get_company_id

router = APIRouter(prefix="/sheets", tags=["Spreadsheet"])


# ── Pydantic schemas ──────────────────────────────────────────

class WorkbookCreate(BaseModel):
    name: str

class WorkbookRename(BaseModel):
    name: str

class SheetCreate(BaseModel):
    name: str

class SheetRename(BaseModel):
    name: str
    sort_order: Optional[int] = None

class ColumnCreate(BaseModel):
    name: str
    width: Optional[int] = 120
    col_type: Optional[str] = "text"   # text | number | date | select

class ColumnUpdate(BaseModel):
    name: Optional[str] = None
    width: Optional[int] = None
    sort_order: Optional[int] = None
    col_type: Optional[str] = None

class RowAdd(BaseModel):
    count: int = 1           # how many blank rows to add
    after_row_id: Optional[int] = None   # insert after this row (None = append)

class RowColour(BaseModel):
    colour: Optional[str] = None   # hex or None to clear

class CellUpdate(BaseModel):
    row_id:      int
    col_id:      int
    value:       Optional[str]  = None
    # formatting (None = don't change)
    bold:        Optional[bool] = None
    italic:      Optional[bool] = None
    underline:   Optional[bool] = None
    strike:      Optional[bool] = None
    font_size:   Optional[int]  = None
    font_colour: Optional[str]  = None   # hex or "" to clear
    fill_colour: Optional[str]  = None   # hex or "" to clear
    align:       Optional[str]  = None   # left|center|right
    wrap:        Optional[bool] = None
    border:      Optional[str]  = None   # none|all|outer|bottom

class BulkCellUpdate(BaseModel):
    cells: List[CellUpdate]
    updated_by: Optional[str] = None


# ── Helpers ───────────────────────────────────────────────────

def _cell_fmt(cell) -> dict:
    """Serialise a cell with its value and formatting."""
    return {
        "value":       cell.value or "",
        "bold":        bool(cell.bold),
        "italic":      bool(cell.italic),
        "underline":   bool(cell.underline),
        "strike":      bool(cell.strike),
        "font_size":   cell.font_size or 13,
        "font_colour": cell.font_colour or "",
        "fill_colour": cell.fill_colour or "",
        "align":       cell.align or "left",
        "wrap":        bool(cell.wrap),
        "border":      cell.border or "none",
    }

def _sheet_to_dict(sheet: SpreadsheetSheet, db: Session) -> dict:
    """Serialise a sheet with its columns, rows, and cells."""
    cols = sorted(sheet.columns, key=lambda c: c.sort_order)
    rows = sorted(sheet.rows,    key=lambda r: r.sort_order)

    # Build a cell lookup: {(row_id, col_id): cell_obj}
    cell_map = {(c.row_id, c.col_id): c for c in sheet.cells}

    rows_data = []
    for row in rows:
        cells = {}
        for col in cols:
            c = cell_map.get((row.id, col.id))
            cells[col.id] = _cell_fmt(c) if c else {
                "value": "", "bold": False, "italic": False, "underline": False,
                "strike": False, "font_size": 13, "font_colour": "", "fill_colour": "",
                "align": "left", "wrap": False, "border": "none",
            }
        rows_data.append({
            "id":      row.id,
            "colour":  row.colour or "",
            "cells":   cells,
        })

    return {
        "id":         sheet.id,
        "name":       sheet.name,
        "sort_order": sheet.sort_order,
        "columns":    [{"id": c.id, "name": c.name, "width": c.width,
                        "col_type": c.col_type, "sort_order": c.sort_order}
                       for c in cols],
        "rows":       rows_data,
    }


def _touch_workbook(wb: SpreadsheetWorkbook):
    wb.updated_at = datetime.utcnow()


# ── Workbooks ─────────────────────────────────────────────────

@router.get("/workbooks/")
def list_workbooks(
    db: Session = Depends(get_db),
    company_id: int = Depends(get_company_id),
):
    wbs = db.query(SpreadsheetWorkbook).filter(
        SpreadsheetWorkbook.company_id == company_id,
    ).order_by(SpreadsheetWorkbook.created_at.asc()).all()
    return [{"id": w.id, "name": w.name, "updated_at": w.updated_at.isoformat() if w.updated_at else None,
             "sheet_count": len(w.sheets)} for w in wbs]


@router.post("/workbooks/")
def create_workbook(
    data: WorkbookCreate,
    db: Session = Depends(get_db),
    company_id: int = Depends(get_company_id),
):
    wb = SpreadsheetWorkbook(name=data.name, updated_at=datetime.utcnow(), company_id=company_id)
    db.add(wb)
    db.flush()  # get wb.id

    # Create a default sheet with default columns
    sheet = SpreadsheetSheet(workbook_id=wb.id, name="Sheet 1", sort_order=0)
    db.add(sheet)
    db.flush()

    default_cols = ["A", "B", "C", "D", "E"]
    for i, col_name in enumerate(default_cols):
        db.add(SpreadsheetColumn(sheet_id=sheet.id, name=col_name, sort_order=i, width=120))

    # Add 20 blank rows
    for i in range(20):
        db.add(SpreadsheetRow(sheet_id=sheet.id, sort_order=i))

    db.commit()
    db.refresh(wb)
    return {"id": wb.id, "name": wb.name}


@router.delete("/workbooks/{wb_id}")
def delete_workbook(
    wb_id: int,
    db: Session = Depends(get_db),
    company_id: int = Depends(get_company_id),
):
    wb = db.query(SpreadsheetWorkbook).filter(
        SpreadsheetWorkbook.id == wb_id,
        SpreadsheetWorkbook.company_id == company_id,
    ).first()
    if not wb:
        raise HTTPException(status_code=404, detail="Workbook not found")
    db.delete(wb)
    db.commit()
    return {"ok": True}


@router.patch("/workbooks/{wb_id}")
def rename_workbook(
    wb_id: int,
    data: WorkbookRename,
    db: Session = Depends(get_db),
    company_id: int = Depends(get_company_id),
):
    wb = db.query(SpreadsheetWorkbook).filter(
        SpreadsheetWorkbook.id == wb_id,
        SpreadsheetWorkbook.company_id == company_id,
    ).first()
    if not wb:
        raise HTTPException(status_code=404, detail="Workbook not found")
    wb.name = data.name
    _touch_workbook(wb)
    db.commit()
    return {"id": wb.id, "name": wb.name}


# ── Full workbook fetch + poll ────────────────────────────────

@router.get("/{workbook_id}/poll")
def poll_workbook(
    workbook_id: int,
    db: Session = Depends(get_db),
    company_id: int = Depends(get_company_id),
):
    wb = db.query(SpreadsheetWorkbook).filter(
        SpreadsheetWorkbook.id == workbook_id,
        SpreadsheetWorkbook.company_id == company_id,
    ).first()
    if not wb:
        raise HTTPException(status_code=404, detail="Workbook not found")
    return {
        "workbook_id": wb.id,
        "last_updated": wb.updated_at.isoformat() if wb.updated_at else None,
    }


@router.get("/{workbook_id}/")
def get_workbook(
    workbook_id: int,
    db: Session = Depends(get_db),
    company_id: int = Depends(get_company_id),
):
    wb = db.query(SpreadsheetWorkbook).filter(
        SpreadsheetWorkbook.id == workbook_id,
        SpreadsheetWorkbook.company_id == company_id,
    ).first()
    if not wb:
        raise HTTPException(status_code=404, detail="Workbook not found")

    sheets = sorted(wb.sheets, key=lambda s: s.sort_order)
    return {
        "id":         wb.id,
        "name":       wb.name,
        "updated_at": wb.updated_at.isoformat() if wb.updated_at else None,
        "sheets":     [_sheet_to_dict(s, db) for s in sheets],
    }


# ── Sheets ────────────────────────────────────────────────────

@router.post("/{workbook_id}/")
def create_sheet(workbook_id: int, data: SheetCreate, db: Session = Depends(get_db), company_id: int = Depends(get_company_id)):
    wb = db.query(SpreadsheetWorkbook).filter(
        SpreadsheetWorkbook.id == workbook_id,
        SpreadsheetWorkbook.company_id == company_id,
    ).first()
    if not wb:
        raise HTTPException(status_code=404, detail="Workbook not found")

    max_order = db.query(func.max(SpreadsheetSheet.sort_order)).filter(
        SpreadsheetSheet.workbook_id == workbook_id).scalar() or -1

    sheet = SpreadsheetSheet(workbook_id=workbook_id, name=data.name, sort_order=max_order + 1)
    db.add(sheet)
    db.flush()

    # Default 5 columns + 20 rows
    for i, col_name in enumerate(["A", "B", "C", "D", "E"]):
        db.add(SpreadsheetColumn(sheet_id=sheet.id, name=col_name, sort_order=i, width=120))
    for i in range(20):
        db.add(SpreadsheetRow(sheet_id=sheet.id, sort_order=i))

    _touch_workbook(wb)
    db.commit()
    db.refresh(sheet)
    return _sheet_to_dict(sheet, db)


@router.delete("/{workbook_id}/{sheet_id}")
def delete_sheet(workbook_id: int, sheet_id: int, db: Session = Depends(get_db)):
    sheet = db.query(SpreadsheetSheet).filter(
        SpreadsheetSheet.id == sheet_id,
        SpreadsheetSheet.workbook_id == workbook_id
    ).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")
    # prevent deleting the last sheet
    count = db.query(func.count(SpreadsheetSheet.id)).filter(
        SpreadsheetSheet.workbook_id == workbook_id).scalar()
    if count <= 1:
        raise HTTPException(status_code=400, detail="Cannot delete the last sheet")

    wb = db.query(SpreadsheetWorkbook).filter(SpreadsheetWorkbook.id == workbook_id).first()
    db.delete(sheet)
    _touch_workbook(wb)
    db.commit()
    return {"ok": True}


@router.patch("/{workbook_id}/{sheet_id}")
def rename_sheet(workbook_id: int, sheet_id: int, data: SheetRename, db: Session = Depends(get_db)):
    sheet = db.query(SpreadsheetSheet).filter(
        SpreadsheetSheet.id == sheet_id,
        SpreadsheetSheet.workbook_id == workbook_id
    ).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")
    sheet.name = data.name
    if data.sort_order is not None:
        sheet.sort_order = data.sort_order
    wb = db.query(SpreadsheetWorkbook).filter(SpreadsheetWorkbook.id == workbook_id).first()
    _touch_workbook(wb)
    db.commit()
    return {"id": sheet.id, "name": sheet.name, "sort_order": sheet.sort_order}


# ── Columns ───────────────────────────────────────────────────

@router.post("/{workbook_id}/{sheet_id}/columns")
def add_column(workbook_id: int, sheet_id: int, data: ColumnCreate, db: Session = Depends(get_db)):
    sheet = db.query(SpreadsheetSheet).filter(
        SpreadsheetSheet.id == sheet_id,
        SpreadsheetSheet.workbook_id == workbook_id
    ).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")

    max_order = db.query(func.max(SpreadsheetColumn.sort_order)).filter(
        SpreadsheetColumn.sheet_id == sheet_id).scalar() or -1

    col = SpreadsheetColumn(
        sheet_id=sheet_id, name=data.name, width=data.width or 120,
        col_type=data.col_type or "text", sort_order=max_order + 1
    )
    db.add(col)
    wb = db.query(SpreadsheetWorkbook).filter(SpreadsheetWorkbook.id == workbook_id).first()
    _touch_workbook(wb)
    db.commit()
    db.refresh(col)
    return {"id": col.id, "name": col.name, "width": col.width, "col_type": col.col_type, "sort_order": col.sort_order}


@router.delete("/{workbook_id}/{sheet_id}/columns/{col_id}")
def delete_column(workbook_id: int, sheet_id: int, col_id: int, db: Session = Depends(get_db)):
    col = db.query(SpreadsheetColumn).filter(
        SpreadsheetColumn.id == col_id,
        SpreadsheetColumn.sheet_id == sheet_id
    ).first()
    if not col:
        raise HTTPException(status_code=404, detail="Column not found")
    # prevent deleting the last column
    count = db.query(func.count(SpreadsheetColumn.id)).filter(
        SpreadsheetColumn.sheet_id == sheet_id).scalar()
    if count <= 1:
        raise HTTPException(status_code=400, detail="Cannot delete the last column")

    db.delete(col)
    wb = db.query(SpreadsheetWorkbook).filter(SpreadsheetWorkbook.id == workbook_id).first()
    _touch_workbook(wb)
    db.commit()
    return {"ok": True}


@router.patch("/{workbook_id}/{sheet_id}/columns/{col_id}")
def update_column(workbook_id: int, sheet_id: int, col_id: int, data: ColumnUpdate, db: Session = Depends(get_db)):
    col = db.query(SpreadsheetColumn).filter(
        SpreadsheetColumn.id == col_id,
        SpreadsheetColumn.sheet_id == sheet_id
    ).first()
    if not col:
        raise HTTPException(status_code=404, detail="Column not found")
    if data.name       is not None: col.name       = data.name
    if data.width      is not None: col.width      = data.width
    if data.sort_order is not None: col.sort_order = data.sort_order
    if data.col_type   is not None: col.col_type   = data.col_type
    wb = db.query(SpreadsheetWorkbook).filter(SpreadsheetWorkbook.id == workbook_id).first()
    _touch_workbook(wb)
    db.commit()
    return {"id": col.id, "name": col.name, "width": col.width, "col_type": col.col_type, "sort_order": col.sort_order}


# ── Rows ──────────────────────────────────────────────────────

@router.post("/{workbook_id}/{sheet_id}/rows")
def add_rows(workbook_id: int, sheet_id: int, data: RowAdd, db: Session = Depends(get_db)):
    sheet = db.query(SpreadsheetSheet).filter(
        SpreadsheetSheet.id == sheet_id,
        SpreadsheetSheet.workbook_id == workbook_id
    ).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")

    max_order = db.query(func.max(SpreadsheetRow.sort_order)).filter(
        SpreadsheetRow.sheet_id == sheet_id).scalar() or -1

    new_rows = []
    for i in range(max(1, min(data.count, 100))):
        row = SpreadsheetRow(sheet_id=sheet_id, sort_order=max_order + 1 + i)
        db.add(row)
        new_rows.append(row)

    wb = db.query(SpreadsheetWorkbook).filter(SpreadsheetWorkbook.id == workbook_id).first()
    _touch_workbook(wb)
    db.commit()
    return [{"id": r.id, "sort_order": r.sort_order, "colour": ""} for r in new_rows]


@router.delete("/{workbook_id}/{sheet_id}/rows/{row_id}")
def delete_row(workbook_id: int, sheet_id: int, row_id: int, db: Session = Depends(get_db)):
    row = db.query(SpreadsheetRow).filter(
        SpreadsheetRow.id == row_id,
        SpreadsheetRow.sheet_id == sheet_id
    ).first()
    if not row:
        raise HTTPException(status_code=404, detail="Row not found")
    db.delete(row)
    wb = db.query(SpreadsheetWorkbook).filter(SpreadsheetWorkbook.id == workbook_id).first()
    _touch_workbook(wb)
    db.commit()
    return {"ok": True}


@router.patch("/{workbook_id}/{sheet_id}/rows/{row_id}")
def update_row(workbook_id: int, sheet_id: int, row_id: int, data: RowColour, db: Session = Depends(get_db)):
    row = db.query(SpreadsheetRow).filter(
        SpreadsheetRow.id == row_id,
        SpreadsheetRow.sheet_id == sheet_id
    ).first()
    if not row:
        raise HTTPException(status_code=404, detail="Row not found")
    row.colour = data.colour or None
    wb = db.query(SpreadsheetWorkbook).filter(SpreadsheetWorkbook.id == workbook_id).first()
    _touch_workbook(wb)
    db.commit()
    return {"id": row.id, "colour": row.colour or ""}


# ── Cells ─────────────────────────────────────────────────────

@router.patch("/{workbook_id}/{sheet_id}/cells")
def update_cells(workbook_id: int, sheet_id: int, data: BulkCellUpdate, db: Session = Depends(get_db)):
    """Upsert multiple cells at once. Blank value = clear the cell."""
    for cu in data.cells:
        # Validate row belongs to this sheet
        row = db.query(SpreadsheetRow).filter(
            SpreadsheetRow.id == cu.row_id,
            SpreadsheetRow.sheet_id == sheet_id
        ).first()
        col = db.query(SpreadsheetColumn).filter(
            SpreadsheetColumn.id == cu.col_id,
            SpreadsheetColumn.sheet_id == sheet_id
        ).first()
        if not row or not col:
            continue

        cell = db.query(SpreadsheetCell).filter(
            SpreadsheetCell.row_id == cu.row_id,
            SpreadsheetCell.col_id == cu.col_id
        ).first()

        # Determine if this cell has any non-default content
        has_value = cu.value is not None and cu.value != ""
        has_fmt = any([
            cu.bold is not None, cu.italic is not None, cu.underline is not None,
            cu.strike is not None, cu.font_size is not None,
            cu.font_colour is not None, cu.fill_colour is not None,
            cu.align is not None, cu.wrap is not None, cu.border is not None,
        ])

        if not has_value and not has_fmt and not cell:
            continue  # nothing to do

        if not cell:
            cell = SpreadsheetCell(
                sheet_id=sheet_id, row_id=cu.row_id, col_id=cu.col_id
            )
            db.add(cell)

        if cu.value       is not None: cell.value       = cu.value
        if cu.bold        is not None: cell.bold        = cu.bold
        if cu.italic      is not None: cell.italic      = cu.italic
        if cu.underline   is not None: cell.underline   = cu.underline
        if cu.strike      is not None: cell.strike      = cu.strike
        if cu.font_size   is not None: cell.font_size   = cu.font_size
        if cu.font_colour is not None: cell.font_colour = cu.font_colour or None
        if cu.fill_colour is not None: cell.fill_colour = cu.fill_colour or None
        if cu.align       is not None: cell.align       = cu.align
        if cu.wrap        is not None: cell.wrap        = cu.wrap
        if cu.border      is not None: cell.border      = cu.border

    wb = db.query(SpreadsheetWorkbook).filter(SpreadsheetWorkbook.id == workbook_id).first()
    if wb:
        _touch_workbook(wb)
    db.commit()
    return {"ok": True, "updated": len(data.cells)}


# ── Excel Export ──────────────────────────────────────────────

@router.get("/{workbook_id}/{sheet_id}/export")
def export_sheet(workbook_id: int, sheet_id: int, db: Session = Depends(get_db)):
    """Export a single sheet as an Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    sheet = db.query(SpreadsheetSheet).filter(
        SpreadsheetSheet.id == sheet_id,
        SpreadsheetSheet.workbook_id == workbook_id
    ).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")

    wb_obj = openpyxl.Workbook()
    ws = wb_obj.active
    ws.title = sheet.name[:31]  # Excel max sheet name length

    cols = sorted(sheet.columns, key=lambda c: c.sort_order)
    rows = sorted(sheet.rows,    key=lambda r: r.sort_order)
    cell_map = {(c.row_id, c.col_id): c for c in sheet.cells}

    # Header row
    HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col.name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = max((col.width or 120) / 7, 10)
    ws.row_dimensions[1].height = 22

    # Data rows
    for ri, row in enumerate(rows, 2):
        row_colour = row.colour
        for ci, col in enumerate(cols, 1):
            cell_obj = cell_map.get((row.id, col.id))
            value = cell_obj.value if cell_obj else ""
            c = ws.cell(row=ri, column=ci, value=value)
            # Cell fill: cell fill_colour takes priority, then row colour
            fill_hex = (cell_obj.fill_colour or "").lstrip("#") if cell_obj else ""
            row_hex  = (row_colour or "").lstrip("#")
            if fill_hex:
                c.fill = PatternFill("solid", fgColor=fill_hex)
            elif row_hex:
                c.fill = PatternFill("solid", fgColor=row_hex)
            # Font
            font_kw = {"size": (cell_obj.font_size or 13) if cell_obj else 13}
            if cell_obj:
                if cell_obj.bold:      font_kw["bold"]      = True
                if cell_obj.italic:    font_kw["italic"]    = True
                if cell_obj.underline: font_kw["underline"] = "single"
                if cell_obj.strike:    font_kw["strike"]    = True
                if cell_obj.font_colour:
                    font_kw["color"] = cell_obj.font_colour.lstrip("#")
            c.font = Font(**font_kw)
            align_h = (cell_obj.align or "left") if cell_obj else "left"
            wrap    = bool(cell_obj.wrap) if cell_obj else False
            c.alignment = Alignment(horizontal=align_h, vertical="center", wrap_text=wrap)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb_obj.save(buf)
    buf.seek(0)

    safe_name = sheet.name.replace(" ", "_").replace("/", "-")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}.xlsx"'},
    )


@router.get("/{workbook_id}/export-all")
def export_workbook(workbook_id: int, db: Session = Depends(get_db)):
    """Export all sheets as a multi-sheet Excel workbook."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb_rec = db.query(SpreadsheetWorkbook).filter(SpreadsheetWorkbook.id == workbook_id).first()
    if not wb_rec:
        raise HTTPException(status_code=404, detail="Workbook not found")

    wb_obj = openpyxl.Workbook()
    wb_obj.remove(wb_obj.active)  # remove default sheet

    sheets = sorted(wb_rec.sheets, key=lambda s: s.sort_order)

    HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)

    for sheet in sheets:
        ws = wb_obj.create_sheet(title=sheet.name[:31])
        cols = sorted(sheet.columns, key=lambda c: c.sort_order)
        rows = sorted(sheet.rows,    key=lambda r: r.sort_order)
        cell_map = {(c.row_id, c.col_id): c for c in sheet.cells}

        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=col.name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(ci)].width = max((col.width or 120) / 7, 10)
        ws.row_dimensions[1].height = 22

        for ri, row in enumerate(rows, 2):
            row_colour = row.colour
            for ci, col in enumerate(cols, 1):
                cell_obj = cell_map.get((row.id, col.id))
                value = cell_obj.value if cell_obj else ""
                c = ws.cell(row=ri, column=ci, value=value)
                fill_hex = (cell_obj.fill_colour or "").lstrip("#") if cell_obj else ""
                row_hex  = (row_colour or "").lstrip("#")
                if fill_hex:
                    c.fill = PatternFill("solid", fgColor=fill_hex)
                elif row_hex:
                    c.fill = PatternFill("solid", fgColor=row_hex)
                font_kw = {"size": (cell_obj.font_size or 13) if cell_obj else 13}
                if cell_obj:
                    if cell_obj.bold:      font_kw["bold"]      = True
                    if cell_obj.italic:    font_kw["italic"]    = True
                    if cell_obj.underline: font_kw["underline"] = "single"
                    if cell_obj.strike:    font_kw["strike"]    = True
                    if cell_obj.font_colour:
                        font_kw["color"] = cell_obj.font_colour.lstrip("#")
                c.font = Font(**font_kw)
                align_h = (cell_obj.align or "left") if cell_obj else "left"
                wrap    = bool(cell_obj.wrap) if cell_obj else False
                c.alignment = Alignment(horizontal=align_h, vertical="center", wrap_text=wrap)
        ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb_obj.save(buf)
    buf.seek(0)

    safe_name = wb_rec.name.replace(" ", "_").replace("/", "-")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}.xlsx"'},
    )
