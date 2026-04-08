"""
Reports & Exports router
Endpoints:
  GET /reports/inventory-snapshot          → JSON
  GET /reports/inventory-snapshot/export   → Excel download
  GET /reports/dispatch-history            → JSON
  GET /reports/dispatch-history/export     → Excel
  GET /reports/receiving-history           → JSON
  GET /reports/receiving-history/export    → Excel
  GET /reports/low-stock                   → JSON
  GET /reports/low-stock/export            → Excel
  GET /reports/expiry-report               → JSON
  GET /reports/expiry-report/export        → Excel
"""
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import date, timedelta, datetime
from typing import Optional
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import sys, os
sys.path.append(os.path.dirname(os.path.dirname(__file__)))
from database import get_db
from models import SKU, Inventory, Batch, DispatchRecord, DispatchRecordItem, InventoryAdjustment, Invoice, InvoicePayment, PurchaseOrder, Customer
from security import get_current_user, get_company_id
from sqlalchemy import func, extract

router = APIRouter(prefix="/reports", tags=["Reports"])

# ── Excel styling helpers ─────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1E3A5F")   # dark blue
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL     = PatternFill("solid", fgColor="F0F4F8")
BORDER_THIN  = Border(
    bottom=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="E5E7EB"),
)

def _style_header(ws, row_num: int, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN

def _style_row(ws, row_num: int, col_count: int, alt: bool = False):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        if alt:
            cell.fill = ALT_FILL
        cell.alignment = Alignment(vertical="center")
        cell.border = BORDER_THIN

def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)

def _make_excel_response(wb: openpyxl.Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

def _report_title(ws, title: str, subtitle: str, col_count: int):
    """Add a 2-row title block above the data headers."""
    ws.insert_rows(1, 2)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
    t = ws.cell(1, 1, title)
    t.font = Font(bold=True, size=14, color="1E3A5F")
    t.alignment = Alignment(horizontal="left", vertical="center")
    s = ws.cell(2, 1, subtitle)
    s.font = Font(size=10, color="6B7280")
    s.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 16


# ══════════════════════════════════════════════════════════════
# 1. INVENTORY SNAPSHOT
# ══════════════════════════════════════════════════════════════
def _inventory_snapshot_data(db: Session, company_id: int = None):
    today = date.today()
    q = db.query(SKU).filter(SKU.is_active == True)
    if company_id is not None:
        q = q.filter(SKU.company_id == company_id)
    skus = q.order_by(SKU.category, SKU.product_name).all()
    rows = []
    for sku in skus:
        inv_q = db.query(Inventory).filter(Inventory.sku_id == sku.id)
        if company_id is not None:
            inv_q = inv_q.filter(Inventory.company_id == company_id)
        inv_rows = inv_q.all()
        wh_map = {i.warehouse: i.cases_on_hand for i in inv_rows}
        wh1 = wh_map.get("WH1", 0)
        wh2 = wh_map.get("WH2", 0)
        total = wh1 + wh2

        # Earliest expiry across all batches
        batch_q = db.query(Batch).filter(
            Batch.sku_id == sku.id,
            Batch.cases_remaining > 0,
            Batch.has_expiry == True,
            Batch.expiry_date != None,
        )
        if company_id is not None:
            batch_q = batch_q.filter(Batch.company_id == company_id)
        earliest_batch = batch_q.order_by(Batch.expiry_date.asc()).first()

        earliest_expiry = earliest_batch.expiry_date if earliest_batch else None
        days_to_expiry  = (earliest_expiry - today).days if earliest_expiry else None

        if total == 0:
            status = "Stockout"
        elif total <= sku.reorder_point:
            status = "Low Stock"
        elif sku.max_stock and total > sku.max_stock:
            status = "Overstock"
        else:
            status = "OK"

        rows.append({
            "sku_code":       sku.sku_code,
            "product_name":   sku.product_name,
            "category":       sku.category,
            "case_contents":  f"{sku.case_size} {sku.unit_label}",
            "wh1_cases":      wh1,
            "wh2_cases":      wh2,
            "total_cases":    total,
            "reorder_point":  sku.reorder_point,
            "max_stock":      sku.max_stock,
            "status":         status,
            "earliest_expiry": earliest_expiry.isoformat() if earliest_expiry else "—",
            "days_to_expiry":  days_to_expiry,
        })
    return rows

@router.get("/inventory-snapshot")
def inventory_snapshot(
    db: Session = Depends(get_db),
    company_id: int = Depends(get_company_id),
):
    return {"as_of": date.today().isoformat(), "rows": _inventory_snapshot_data(db, company_id)}

@router.get("/inventory-snapshot/export")
def inventory_snapshot_export(
    db: Session = Depends(get_db),
    company_id: int = Depends(get_company_id),
):
    rows = _inventory_snapshot_data(db, company_id)
    today = date.today()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Snapshot"

    headers = ["SKU Code","Product Name","Category","Case Contents",
               "WH1 Cases","WH2 Cases","Total Cases","Reorder At","Max Stock","Status","Earliest Expiry","Days to Expiry"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    for i, r in enumerate(rows):
        ws.append([
            r["sku_code"], r["product_name"], r["category"], r["case_contents"],
            r["wh1_cases"], r["wh2_cases"], r["total_cases"],
            r["reorder_point"], r["max_stock"], r["status"],
            r["earliest_expiry"], r["days_to_expiry"] if r["days_to_expiry"] is not None else "—",
        ])
        _style_row(ws, i + 2, len(headers), alt=(i % 2 == 1))
        # Colour status cell
        status_cell = ws.cell(i + 2, 10)
        if r["status"] == "Stockout":
            status_cell.fill = PatternFill("solid", fgColor="FEE2E2")
            status_cell.font = Font(bold=True, color="DC2626")
        elif r["status"] == "Low Stock":
            status_cell.fill = PatternFill("solid", fgColor="FEF3C7")
            status_cell.font = Font(bold=True, color="D97706")

    _report_title(ws, "Inventory Snapshot", f"Generated: {today.strftime('%d %b %Y')}", len(headers))
    ws.freeze_panes = "A4"
    _auto_width(ws)

    return _make_excel_response(wb, f"inventory_snapshot_{today.isoformat()}.xlsx")


# ══════════════════════════════════════════════════════════════
# 2. DISPATCH HISTORY
# ══════════════════════════════════════════════════════════════
def _dispatch_history_data(db: Session, date_from: Optional[date] = None, date_to: Optional[date] = None, company_id: int = None):
    q = db.query(DispatchRecordItem).join(
        DispatchRecord, DispatchRecordItem.dispatch_id == DispatchRecord.id
    )
    if company_id is not None:
        q = q.filter(DispatchRecord.company_id == company_id)
    if date_from:
        q = q.filter(DispatchRecord.dispatch_date >= date_from)
    if date_to:
        q = q.filter(DispatchRecord.dispatch_date <= date_to)
    q = q.order_by(DispatchRecord.dispatch_date.desc(), DispatchRecord.id.desc())

    rows = []
    for item in q.all():
        rows.append({
            "dispatch_date":   item.dispatch.dispatch_date.isoformat(),
            "ref":             item.dispatch.ref,
            "note":            item.dispatch.note or "",
            "sku_code":        item.sku.sku_code,
            "product_name":    item.sku.product_name,
            "cases_requested": item.cases_requested,
            "cases_fulfilled": item.cases_fulfilled,
            "shortfall":       item.cases_requested - item.cases_fulfilled,
        })
    return rows

@router.get("/dispatch-history")
def dispatch_history(
    date_from: Optional[date] = None,
    date_to:   Optional[date] = None,
    db: Session = Depends(get_db),
    company_id: int = Depends(get_company_id),
):
    return {"rows": _dispatch_history_data(db, date_from, date_to, company_id)}

@router.get("/dispatch-history/export")
def dispatch_history_export(
    date_from: Optional[date] = None,
    date_to:   Optional[date] = None,
    db: Session = Depends(get_db),
    company_id: int = Depends(get_company_id),
):
    rows = _dispatch_history_data(db, date_from, date_to, company_id)
    today = date.today()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dispatch History"

    headers = ["Date","Reference","Note","SKU Code","Product Name",
               "Cases Requested","Cases Fulfilled","Shortfall"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    for i, r in enumerate(rows):
        ws.append([
            r["dispatch_date"], r["ref"], r["note"],
            r["sku_code"], r["product_name"],
            r["cases_requested"], r["cases_fulfilled"], r["shortfall"],
        ])
        _style_row(ws, i + 2, len(headers), alt=(i % 2 == 1))
        if r["shortfall"] > 0:
            ws.cell(i + 2, 8).fill = PatternFill("solid", fgColor="FEE2E2")

    subtitle = f"Generated: {today.strftime('%d %b %Y')}"
    if date_from or date_to:
        subtitle += f"  |  Period: {date_from or 'start'} → {date_to or 'today'}"
    _report_title(ws, "Dispatch History", subtitle, len(headers))
    ws.freeze_panes = "A4"
    _auto_width(ws)

    return _make_excel_response(wb, f"dispatch_history_{today.isoformat()}.xlsx")


# ══════════════════════════════════════════════════════════════
# 3. RECEIVING HISTORY
# ══════════════════════════════════════════════════════════════
def _receiving_history_data(db: Session, date_from: Optional[date] = None, date_to: Optional[date] = None, company_id: int = None):
    q = db.query(Batch)
    if company_id is not None:
        q = q.filter(Batch.company_id == company_id)
    if date_from:
        q = q.filter(Batch.received_date >= date_from)
    if date_to:
        q = q.filter(Batch.received_date <= date_to)
    q = q.order_by(Batch.received_date.desc(), Batch.id.desc())

    today = date.today()
    rows = []
    for b in q.all():
        days_to_exp = None
        if b.expiry_date:
            days_to_exp = (b.expiry_date - today).days
        rows.append({
            "received_date":   b.received_date.isoformat(),
            "batch_code":      b.batch_code,
            "sku_code":        b.sku.sku_code,
            "product_name":    b.sku.product_name,
            "category":        b.sku.category,
            "cases_received":  b.cases_received,
            "cases_remaining": b.cases_remaining,
            "warehouse":       b.warehouse,
            "expiry_date":     b.expiry_date.isoformat() if b.expiry_date else "—",
            "days_to_expiry":  days_to_exp,
            "supplier_ref":    b.supplier_ref or "",
            "notes":           b.notes or "",
        })
    return rows

@router.get("/receiving-history")
def receiving_history(
    date_from: Optional[date] = None,
    date_to:   Optional[date] = None,
    db: Session = Depends(get_db),
    company_id: int = Depends(get_company_id),
):
    return {"rows": _receiving_history_data(db, date_from, date_to, company_id)}

@router.get("/receiving-history/export")
def receiving_history_export(
    date_from: Optional[date] = None,
    date_to:   Optional[date] = None,
    db: Session = Depends(get_db),
    company_id: int = Depends(get_company_id),
):
    rows = _receiving_history_data(db, date_from, date_to, company_id)
    today = date.today()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Receiving History"

    headers = ["Date Received","Batch Code","SKU Code","Product Name","Category",
               "Cases Received","Cases Remaining","Warehouse","Expiry Date","Days to Expiry",
               "Supplier Ref","Notes"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    for i, r in enumerate(rows):
        ws.append([
            r["received_date"], r["batch_code"], r["sku_code"], r["product_name"],
            r["category"], r["cases_received"], r["cases_remaining"], r["warehouse"],
            r["expiry_date"],
            r["days_to_expiry"] if r["days_to_expiry"] is not None else "—",
            r["supplier_ref"], r["notes"],
        ])
        _style_row(ws, i + 2, len(headers), alt=(i % 2 == 1))
        # Flag expired or close-to-expiry
        if r["days_to_expiry"] is not None:
            exp_cell = ws.cell(i + 2, 10)
            if r["days_to_expiry"] < 0:
                exp_cell.fill = PatternFill("solid", fgColor="FEE2E2")
                exp_cell.font = Font(bold=True, color="DC2626")
            elif r["days_to_expiry"] <= 30:
                exp_cell.fill = PatternFill("solid", fgColor="FEF3C7")
                exp_cell.font = Font(bold=True, color="D97706")

    subtitle = f"Generated: {today.strftime('%d %b %Y')}"
    if date_from or date_to:
        subtitle += f"  |  Period: {date_from or 'start'} → {date_to or 'today'}"
    _report_title(ws, "Receiving / GRN History", subtitle, len(headers))
    ws.freeze_panes = "A4"
    _auto_width(ws)

    return _make_excel_response(wb, f"receiving_history_{today.isoformat()}.xlsx")


# ══════════════════════════════════════════════════════════════
# 4. LOW STOCK / REORDER REPORT
# ══════════════════════════════════════════════════════════════
def _low_stock_data(db: Session, company_id: int = None):
    q = db.query(SKU).filter(SKU.is_active == True)
    if company_id is not None:
        q = q.filter(SKU.company_id == company_id)
    skus = q.order_by(SKU.category, SKU.product_name).all()
    rows = []
    for sku in skus:
        inv_q = db.query(Inventory).filter(Inventory.sku_id == sku.id)
        if company_id is not None:
            inv_q = inv_q.filter(Inventory.company_id == company_id)
        inv_rows = inv_q.all()
        total = sum(i.cases_on_hand for i in inv_rows)
        if total <= sku.reorder_point:
            rows.append({
                "sku_code":      sku.sku_code,
                "product_name":  sku.product_name,
                "category":      sku.category,
                "total_cases":   total,
                "reorder_point": sku.reorder_point,
                "reorder_qty":   sku.reorder_qty,
                "lead_time_days": sku.lead_time_days,
                "vendor_name":   sku.vendor.name if sku.vendor else "—",
                "status":        "Stockout" if total == 0 else "Low Stock",
                "shortfall":     max(0, sku.reorder_point - total),
            })
    return rows

@router.get("/low-stock")
def low_stock(
    db: Session = Depends(get_db),
    company_id: int = Depends(get_company_id),
):
    return {"rows": _low_stock_data(db, company_id)}

@router.get("/low-stock/export")
def low_stock_export(
    db: Session = Depends(get_db),
    company_id: int = Depends(get_company_id),
):
    rows = _low_stock_data(db, company_id)
    today = date.today()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Low Stock & Reorder"

    headers = ["SKU Code","Product Name","Category","Current Cases",
               "Reorder At","Order Qty","Lead Time (days)","Vendor","Status","Shortfall"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    for i, r in enumerate(rows):
        ws.append([
            r["sku_code"], r["product_name"], r["category"],
            r["total_cases"], r["reorder_point"], r["reorder_qty"],
            r["lead_time_days"], r["vendor_name"], r["status"], r["shortfall"],
        ])
        _style_row(ws, i + 2, len(headers), alt=(i % 2 == 1))
        status_cell = ws.cell(i + 2, 9)
        if r["status"] == "Stockout":
            status_cell.fill = PatternFill("solid", fgColor="FEE2E2")
            status_cell.font = Font(bold=True, color="DC2626")
        else:
            status_cell.fill = PatternFill("solid", fgColor="FEF3C7")
            status_cell.font = Font(bold=True, color="D97706")

    _report_title(ws, "Low Stock & Reorder Report", f"Generated: {today.strftime('%d %b %Y')}  |  {len(rows)} SKUs need attention", len(headers))
    ws.freeze_panes = "A4"
    _auto_width(ws)

    return _make_excel_response(wb, f"low_stock_{today.isoformat()}.xlsx")


# ══════════════════════════════════════════════════════════════
# 5. EXPIRY REPORT
# ══════════════════════════════════════════════════════════════
def _expiry_report_data(db: Session, within_days: int = 90, company_id: int = None):
    today = date.today()
    cutoff = today + timedelta(days=within_days)

    q = db.query(Batch).filter(
        Batch.cases_remaining > 0,
        Batch.has_expiry == True,
        Batch.expiry_date != None,
        Batch.expiry_date <= cutoff,
    )
    if company_id is not None:
        q = q.filter(Batch.company_id == company_id)
    batches = q.order_by(Batch.expiry_date.asc()).all()

    rows = []
    for b in batches:
        days = (b.expiry_date - today).days
        if days < 0:
            urgency = "EXPIRED"
        elif days <= 30:
            urgency = "CRITICAL"
        elif days <= 60:
            urgency = "WARNING"
        else:
            urgency = "MONITOR"
        rows.append({
            "expiry_date":     b.expiry_date.isoformat(),
            "days_to_expiry":  days,
            "urgency":         urgency,
            "batch_code":      b.batch_code,
            "sku_code":        b.sku.sku_code,
            "product_name":    b.sku.product_name,
            "category":        b.sku.category,
            "cases_remaining": b.cases_remaining,
            "warehouse":       b.warehouse,
            "received_date":   b.received_date.isoformat(),
            "supplier_ref":    b.supplier_ref or "",
        })
    return rows

@router.get("/expiry-report")
def expiry_report(
    within_days: int = 90,
    db: Session = Depends(get_db),
    company_id: int = Depends(get_company_id),
):
    rows = _expiry_report_data(db, within_days, company_id)
    today = date.today()
    return {
        "as_of":     today.isoformat(),
        "within_days": within_days,
        "expired_count":  sum(1 for r in rows if r["urgency"] == "EXPIRED"),
        "critical_count": sum(1 for r in rows if r["urgency"] == "CRITICAL"),
        "warning_count":  sum(1 for r in rows if r["urgency"] == "WARNING"),
        "monitor_count":  sum(1 for r in rows if r["urgency"] == "MONITOR"),
        "rows": rows,
    }

@router.get("/expiry-report/export")
def expiry_report_export(
    within_days: int = 90,
    db: Session = Depends(get_db),
    company_id: int = Depends(get_company_id),
):
    rows = _expiry_report_data(db, within_days, company_id)
    today = date.today()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expiry Report"

    headers = ["Expiry Date","Days Left","Urgency","Batch Code","SKU Code","Product Name",
               "Category","Cases Remaining","Warehouse","Received Date","Supplier Ref"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    urgency_colors = {
        "EXPIRED":  ("FEE2E2", "DC2626"),
        "CRITICAL": ("FEF3C7", "B45309"),
        "WARNING":  ("FEF9C3", "CA8A04"),
        "MONITOR":  ("ECFDF5", "059669"),
    }

    for i, r in enumerate(rows):
        ws.append([
            r["expiry_date"], r["days_to_expiry"], r["urgency"],
            r["batch_code"], r["sku_code"], r["product_name"],
            r["category"], r["cases_remaining"], r["warehouse"],
            r["received_date"], r["supplier_ref"],
        ])
        _style_row(ws, i + 2, len(headers), alt=(i % 2 == 1))
        bg, fg = urgency_colors.get(r["urgency"], ("FFFFFF", "000000"))
        urgency_cell = ws.cell(i + 2, 3)
        urgency_cell.fill = PatternFill("solid", fgColor=bg)
        urgency_cell.font = Font(bold=True, color=fg)

    _report_title(ws, f"Expiry Report — Next {within_days} Days",
                  f"Generated: {today.strftime('%d %b %Y')}  |  {len(rows)} batches", len(headers))
    ws.freeze_panes = "A4"
    _auto_width(ws)

    return _make_excel_response(wb, f"expiry_report_{today.isoformat()}.xlsx")


# ══════════════════════════════════════════════════════════════
# 6. ADJUSTMENT / WRITE-OFF HISTORY
# ══════════════════════════════════════════════════════════════
@router.get("/adjustments/export")
def adjustments_export(
    db: Session = Depends(get_db),
    company_id: int = Depends(get_company_id),
):
    adjs = db.query(InventoryAdjustment).filter(
        InventoryAdjustment.company_id == company_id,
    ).order_by(
        InventoryAdjustment.adjusted_at.desc()
    ).all()
    today = date.today()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Adjustments"

    headers = ["Date","SKU ID","Warehouse","Before Qty","After Qty","Change","Reason","Notes"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    for i, a in enumerate(adjs):
        sku = db.query(SKU).filter(SKU.id == a.sku_id, SKU.company_id == company_id).first()
        ws.append([
            a.adjusted_at.strftime('%Y-%m-%d %H:%M'),
            f"{sku.sku_code} — {sku.product_name}" if sku else str(a.sku_id),
            a.warehouse,
            a.before_qty, a.after_qty, a.delta,
            a.reason or "", a.notes or "",
        ])
        _style_row(ws, i + 2, len(headers), alt=(i % 2 == 1))
        delta_cell = ws.cell(i + 2, 6)
        if a.delta < 0:
            delta_cell.font = Font(bold=True, color="DC2626")
        elif a.delta > 0:
            delta_cell.font = Font(bold=True, color="059669")

    _report_title(ws, "Stock Adjustments & Write-offs",
                  f"Generated: {today.strftime('%d %b %Y')}  |  {len(adjs)} records", len(headers))
    ws.freeze_panes = "A4"
    _auto_width(ws)

    return _make_excel_response(wb, f"adjustments_{today.isoformat()}.xlsx")


# ─── Financial / P&L Reports (QuickBooks-style) ────────────────

@router.get("/financials")
def financials(
    months: int = 6,
    db: Session = Depends(get_db),
    company_id: int = Depends(get_company_id),
):
    """
    P&L summary: revenue, COGS (from PO receipts), gross profit by month.
    Also returns AR aging totals and top 5 customers by revenue.
    """
    today = date.today()
    period_start = date(today.year, today.month, 1) - timedelta(days=(months - 1) * 30)

    # ── Monthly Revenue (from invoices that are Paid or Partial) ──
    inv_rows = (
        db.query(Invoice)
        .filter(
            Invoice.company_id == company_id,
            Invoice.status.in_(["Paid", "Partial", "Sent", "Overdue"]),
            Invoice.invoice_date >= period_start,
        )
        .all()
    )

    # ── Monthly Payments received ──
    pmt_rows = (
        db.query(InvoicePayment)
        .filter(
            InvoicePayment.company_id == company_id,
            InvoicePayment.payment_date >= period_start,
        )
        .all()
    )

    # ── Monthly COGS (from PO receiving cost) ──
    po_rows = (
        db.query(PurchaseOrder)
        .filter(
            PurchaseOrder.company_id == company_id,
            PurchaseOrder.status.in_(["received", "partial"]),
            PurchaseOrder.expected_date >= period_start,
        )
        .all()
    )

    # Build month buckets
    def month_key(d):
        return f"{d.year}-{str(d.month).zfill(2)}"

    # Revenue billed per month
    revenue_by_month: dict = {}
    for inv in inv_rows:
        k = month_key(inv.invoice_date)
        revenue_by_month[k] = revenue_by_month.get(k, 0.0) + (inv.total or 0.0)

    # Cash received per month
    cash_by_month: dict = {}
    for pmt in pmt_rows:
        k = month_key(pmt.payment_date)
        cash_by_month[k] = cash_by_month.get(k, 0.0) + pmt.amount

    # COGS per month (sum of PO item unit_cost * cases_received)
    cogs_by_month: dict = {}
    for po in po_rows:
        k = month_key(po.expected_date or today)
        cogs = sum((it.unit_cost or 0) * it.cases_received for it in po.items)
        cogs_by_month[k] = cogs_by_month.get(k, 0.0) + cogs

    # Build sorted months list
    all_months = sorted(set(
        list(revenue_by_month.keys()) +
        list(cash_by_month.keys()) +
        list(cogs_by_month.keys())
    ))

    monthly = []
    for k in all_months:
        rev   = round(revenue_by_month.get(k, 0.0), 2)
        cash  = round(cash_by_month.get(k, 0.0), 2)
        cogs  = round(cogs_by_month.get(k, 0.0), 2)
        gp    = round(rev - cogs, 2)
        gp_pct = round((gp / rev * 100) if rev > 0 else 0.0, 1)
        monthly.append({"month": k, "revenue": rev, "cash_received": cash,
                         "cogs": cogs, "gross_profit": gp, "gp_pct": gp_pct})

    # ── Totals ───────────────────────────────────────────────
    total_revenue      = round(sum(m["revenue"]       for m in monthly), 2)
    total_cash         = round(sum(m["cash_received"]  for m in monthly), 2)
    total_cogs         = round(sum(m["cogs"]           for m in monthly), 2)
    total_gp           = round(total_revenue - total_cogs, 2)
    total_gp_pct       = round((total_gp / total_revenue * 100) if total_revenue > 0 else 0.0, 1)

    # ── AR outstanding ───────────────────────────────────────
    outstanding = (
        db.query(func.sum(Invoice.grand_total))
        .filter(Invoice.company_id == company_id, Invoice.status.in_(["Sent", "Overdue", "Partial"]))
        .scalar() or 0.0
    )
    overdue = (
        db.query(func.sum(Invoice.grand_total))
        .filter(Invoice.company_id == company_id, Invoice.status == "Overdue")
        .scalar() or 0.0
    )

    # ── Top customers by revenue (all time) ──────────────────
    top_customers = []
    cust_rows = (
        db.query(Invoice.customer_id, func.sum(Invoice.total).label("total_revenue"))
        .filter(Invoice.company_id == company_id, Invoice.status.in_(["Paid", "Partial", "Sent", "Overdue"]))
        .group_by(Invoice.customer_id)
        .order_by(func.sum(Invoice.total).desc())
        .limit(8)
        .all()
    )
    for cid, rev in cust_rows:
        cust = db.query(Customer).filter(Customer.id == cid).first() if cid else None
        top_customers.append({
            "customer_id":   cid,
            "customer_name": cust.name if cust else "Walk-in",
            "revenue":       round(float(rev or 0), 2),
        })

    # ── Payment method breakdown ──────────────────────────────
    method_rows = (
        db.query(InvoicePayment.method, func.sum(InvoicePayment.amount))
        .filter(InvoicePayment.company_id == company_id)
        .group_by(InvoicePayment.method)
        .all()
    )
    payment_methods = [{"method": m, "total": round(float(t or 0), 2)} for m, t in method_rows]

    return {
        "monthly":          monthly,
        "totals": {
            "revenue":      total_revenue,
            "cash_received": total_cash,
            "cogs":         total_cogs,
            "gross_profit": total_gp,
            "gp_pct":       total_gp_pct,
            "outstanding":  round(float(outstanding), 2),
            "overdue":      round(float(overdue), 2),
        },
        "top_customers":    top_customers,
        "payment_methods":  payment_methods,
    }
